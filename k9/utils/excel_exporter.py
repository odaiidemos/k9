"""
Excel Export Utilities for K9 Operations Management System

This module provides reusable functions for exporting data to Excel format (XLSX)
using the openpyxl library. It replaces CSV exports with properly formatted Excel files.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
from typing import List, Any, Optional


def create_excel_file(
    headers: List[str],
    data: List[List[Any]],
    sheet_name: str = "Sheet1",
    title: Optional[str] = None,
    auto_filter: bool = True,
    freeze_header: bool = True
) -> Workbook:
    """
    Create an Excel workbook with formatted data
    
    Args:
        headers: List of column headers
        data: List of rows, where each row is a list of cell values
        sheet_name: Name for the worksheet
        title: Optional title to add above the headers
        auto_filter: Whether to enable auto-filter on headers
        freeze_header: Whether to freeze the header row
        
    Returns:
        openpyxl Workbook object
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    title_font = Font(bold=True, size=14, color="1F4788")
    title_alignment = Alignment(horizontal="center", vertical="center")
    
    cell_alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
    
    border_style = Side(style='thin', color='000000')
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    current_row = 1
    
    # Add title if provided
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        current_row = 2
    
    # Add headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Add data rows
    for row_num, row_data in enumerate(data, current_row + 1):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = cell_alignment
            cell.border = border
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        
        # Check header length
        header_length = len(str(headers[col_num - 1]))
        max_length = max(max_length, header_length)
        
        # Check data lengths
        for row in data:
            if col_num - 1 < len(row):
                cell_value = str(row[col_num - 1]) if row[col_num - 1] is not None else ""
                max_length = max(max_length, len(cell_value))
        
        # Set column width (add some padding)
        adjusted_width = min(max_length + 4, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Enable auto-filter
    if auto_filter and len(headers) > 0:
        ws.auto_filter.ref = ws.dimensions
    
    # Freeze header row
    if freeze_header:
        ws.freeze_panes = ws.cell(row=current_row + 1, column=1)
    
    return wb


def save_excel_to_bytes(workbook: Workbook) -> bytes:
    """
    Save workbook to bytes for download
    
    Args:
        workbook: openpyxl Workbook object
        
    Returns:
        bytes: Excel file content as bytes
    """
    from io import BytesIO
    
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def create_attendance_report_excel(
    records: List[Any],
    start_date: datetime,
    end_date: datetime,
    shift_name: Optional[str] = None
) -> Workbook:
    """
    Create an Excel file for attendance report
    
    Args:
        records: List of attendance records
        start_date: Start date of the report
        end_date: End date of the report
        shift_name: Optional shift name for the report title
        
    Returns:
        openpyxl Workbook object
    """
    # Prepare title
    if shift_name:
        title = f"تقرير حضور - {shift_name} - من {start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')}"
    else:
        title = f"تقرير حضور شامل - من {start_date.strftime('%Y-%m-%d')} إلى {end_date.strftime('%Y-%m-%d')}"
    
    # Define headers
    headers = [
        'التاريخ', 'الوردية', 'نوع العضو', 'رمز العضو', 'اسم العضو', 
        'الحالة', 'سبب الغياب', 'سبب التأخير', 'وقت الدخول', 'وقت الخروج', 'ملاحظات'
    ]
    
    # Prepare data rows
    data = []
    for record in records:
        from k9.models.models import EntityType
        
        row = [
            record.date.strftime('%Y-%m-%d'),
            record.shift.name,
            'موظف' if record.entity_type == EntityType.EMPLOYEE else 'كلب',
            record.get_entity_code(),
            record.get_entity_name(),
            record.get_status_display(),
            record.get_absence_reason_display(),
            record.late_reason or '',
            record.check_in_time.strftime('%H:%M') if record.check_in_time else '',
            record.check_out_time.strftime('%H:%M') if record.check_out_time else '',
            record.notes or ''
        ]
        data.append(row)
    
    return create_excel_file(headers, data, sheet_name="تقرير الحضور", title=title)


def create_permissions_report_excel(permissions: List[Any]) -> Workbook:
    """
    Create an Excel file for permissions export
    
    Args:
        permissions: List of permission records
        
    Returns:
        openpyxl Workbook object
    """
    headers = [
        'User ID', 'Username', 'Full Name', 'Project ID', 'Section', 'Subsection', 
        'Permission Type', 'Is Granted', 'Updated At'
    ]
    
    data = []
    for perm in permissions:
        row = [
            perm.user_id,
            perm.user.username,
            perm.user.full_name,
            perm.project_id or 'Global',
            perm.section,
            perm.subsection,
            perm.permission_type.value,
            'Yes' if perm.is_granted else 'No',
            perm.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ]
        data.append(row)
    
    title = f"All Permissions Export - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    return create_excel_file(headers, data, sheet_name="Permissions", title=title)
