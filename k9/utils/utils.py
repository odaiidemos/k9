import os
from flask import request, current_app
from k9.models.models import AuditLog
from app import db
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from datetime import datetime
import uuid
import arabic_reshaper
from bidi.algorithm import get_display
import re

ALLOWED_EXTENSIONS = {'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'doc', 'docx'}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def log_audit(user_id, action, target_type, target_id, description=None, old_values=None, new_values=None):
    """Log an audit trail entry"""
    import json
    try:
        audit_log = AuditLog()
        audit_log.user_id = user_id
        audit_log.action = action
        audit_log.target_type = target_type
        # Handle target_id properly for UUID compatibility
        # Don't convert User IDs (integers) to UUIDs, leave them as None for User targets
        if target_type == 'User' and isinstance(target_id, (int, str)) and str(target_id).isdigit():
            audit_log.target_id = None  # Don't try to store integer user IDs as UUIDs
        else:
            audit_log.target_id = str(target_id) if target_id is not None else None
        
        # Ensure description is a string, not dict (for SQLite compatibility)
        if isinstance(description, dict):
            audit_log.description = json.dumps(description, ensure_ascii=False)
        else:
            audit_log.description = description
            
        # Ensure old_values and new_values are properly handled for SQLite
        audit_log.old_values = json.dumps(old_values, ensure_ascii=False) if old_values and isinstance(old_values, dict) else old_values
        audit_log.new_values = json.dumps(new_values, ensure_ascii=False) if new_values and isinstance(new_values, dict) else new_values
        
        audit_log.ip_address = request.remote_addr if request else None
        audit_log.user_agent = request.headers.get('User-Agent') if request else None
        db.session.add(audit_log)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print(f"Error logging audit: {str(e)}")

def generate_pdf_report(report_type, start_date, end_date, user, filters=None):
    """
    Build a professional Arabic PDF report.
    :param report_type: one of 'dogs', 'employees', 'training', 'veterinary', 'breeding', 'projects'
    :param start_date: optional start date (datetime.date)
    :param end_date: optional end date (datetime.date)
    :param user: current_user for permission filtering
    :param filters: optional dict with keys appropriate to report_type
    :return: filename of generated PDF
    """
    from k9.models.models import Dog, Employee, TrainingSession, VeterinaryVisit, ProductionCycle, Project
    filters = filters or {}

    # Try to register an Arabic-capable font, fall back to default if not available
    arabic_font = 'Helvetica'  # Default fallback
    
    # Register fonts with fallbacks
    font_files = {
        'Amiri': 'static/fonts/Amiri-Regular.ttf',
        'NotoSansArabic': 'static/fonts/NotoSansArabic-Regular.ttf',
        'DejaVuSans': 'static/fonts/DejaVuSans.ttf'
    }
    
    registered_fonts = []
    
    for font_name, font_file in font_files.items():
        try:
            font_path = os.path.join(current_app.root_path, font_file)
            if os.path.exists(font_path):
                # Verify it's a valid font file (not ZIP)
                with open(font_path, 'rb') as f:
                    header = f.read(4)
                    if header[:2] == b'PK':  # ZIP file signature
                        print(f"Font file {font_file} is corrupted (ZIP file)")
                        continue
                
                pdfmetrics.registerFont(TTFont(font_name, font_path))
                registered_fonts.append(font_name)
                print(f"Successfully registered font: {font_name}")
        except Exception as e:
            print(f"Font registration error for {font_file}: {e}")
            continue
    
    # Set primary Arabic font
    if 'Amiri' in registered_fonts:
        arabic_font = 'Amiri'
    elif 'NotoSansArabic' in registered_fonts:
        arabic_font = 'NotoSansArabic'
    else:
        arabic_font = 'Helvetica'
    
    # Set Latin fallback font  
    latin_font = 'DejaVuSans' if 'DejaVuSans' in registered_fonts else 'Helvetica'
    
    # Initialize Arabic text processing
    ARABIC_RE = re.compile(r'[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]+')
    reshaper = arabic_reshaper.ArabicReshaper({
        'delete_harakat': False,
        'support_ligatures': True,
        'use_unshaped_instead_of_isolated': False
    })
    
    def shape_mixed(txt):
        """Reshape only Arabic segments, preserve Latin text"""
        if not txt:
            return ''
        txt = str(txt)
        def _reshape(m):
            return reshaper.reshape(m.group(0))
        reshaped = ARABIC_RE.sub(_reshape, txt)
        return get_display(reshaped, base_dir='R')
    
    def safe_arabic_text(txt):
        """Safe Arabic text processing for direct canvas drawing"""
        if not txt:
            return ''
        txt = str(txt)
        def _reshape(m):
            return reshaper.reshape(m.group(0))
        reshaped = ARABIC_RE.sub(_reshape, txt)
        return get_display(reshaped, base_dir='R')
    
    # Define paragraph styles
    STYLE_AR_R = ParagraphStyle('AR_R', fontName=arabic_font, fontSize=10, alignment=2)  # RIGHT
    STYLE_LTR = ParagraphStyle('LTR', fontName=arabic_font, fontSize=10, alignment=0)   # LEFT
    STYLE_HDR = ParagraphStyle('HDR', fontName=arabic_font, fontSize=11, alignment=2, textColor=colors.whitesmoke)
    
    def para_ar(v):
        """Create Arabic paragraph with right alignment"""
        return Paragraph(shape_mixed('' if v is None else str(v)), STYLE_AR_R)
    
    def para_ltr(v):
        """Create Latin paragraph with left alignment"""
        return Paragraph('' if v is None else str(v), STYLE_LTR)
    
    def hdr(text_ar):
        """Create header paragraph"""
        return Paragraph(shape_mixed(text_ar), STYLE_HDR)
    
    def build_training_table(training_sessions):
        from reportlab.lib.units import mm
        header = [
            hdr('رقم'),
            hdr('اسم الكلب'),
            hdr('المدرب'),
            hdr('الفئة'),
            hdr('الموضوع'),
            hdr('التاريخ'),
            hdr('المدة (دقيقة)'),
            hdr('التقييم')
        ]
        
        rows = [header]
        for i, session in enumerate(training_sessions, start=1):
            category_map = {
                'OBEDIENCE': 'طاعة', 'DETECTION': 'كشف', 'AGILITY': 'رشاقة',
                'ATTACK': 'هجوم', 'FITNESS': 'لياقة'
            }
            
            rows.append([
                para_ltr(i),
                para_ar(session.dog.name if session.dog else ''),
                para_ar(session.trainer.name if session.trainer else ''),
                para_ar(category_map.get(session.category.value, session.category.value)),
                para_ar(session.subject or ''),
                para_ltr(session.session_date.strftime('%Y-%m-%d') if session.session_date else ''),
                para_ltr(str(session.duration) if session.duration else ''),
                para_ar(f"{session.success_rating}/10" if session.success_rating else '')
            ])
        
        col_widths = [12*mm, 30*mm, 25*mm, 20*mm, 35*mm, 25*mm, 20*mm, 25*mm]
        table = Table(rows, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#305496')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#305496')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,1), (0,-1), 'LEFT'),  # رقم
            ('ALIGN', (5,1), (5,-1), 'LEFT'),  # التاريخ
            ('ALIGN', (6,1), (6,-1), 'LEFT'),  # المدة
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
        ]))
        return table

    # Generate a unique filename and file path
    filename = f"report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)

    # Prepare the document with generous margins (similar to the DOCX templates)
    doc = SimpleDocTemplate(filepath, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)

    story = []
    styles = getSampleStyleSheet()

    # Custom style definitions: Arabic text should be right-aligned and use the registered font
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontName=arabic_font,
                                 fontSize=18, textColor=colors.HexColor('#C00000'),
                                 alignment=1, spaceAfter=20)  # centre-aligned red title
    header_style = ParagraphStyle('Header', parent=styles['Heading2'], fontName=arabic_font,
                                  fontSize=12, alignment=2)  # right-aligned
    normal_style = ParagraphStyle('NormalAr', parent=styles['Normal'], fontName=arabic_font,
                                  fontSize=10, alignment=2)

    # Draw header: logo (top left), title (centre), day/date (top right)
    def build_header(canvas_obj, doc_obj):
        width, height = A4
        # Company logo; put a PNG named 'logo.png' inside static/img or adjust path accordingly
        logo_path = os.path.join(current_app.root_path, 'static/img/logo.png')
        if os.path.exists(logo_path):
            canvas_obj.drawImage(logo_path, x=doc_obj.leftMargin,
                                 y=height - doc_obj.topMargin + 20, width=50, height=50, preserveAspectRatio=True)
        # Day/Date placeholders on the right
        canvas_obj.setFont(arabic_font, 12)
        canvas_obj.drawRightString(width - doc_obj.rightMargin, height - doc_obj.topMargin + 50,
                                   safe_arabic_text("اليوم:               "))
        canvas_obj.drawRightString(width - doc_obj.rightMargin, height - doc_obj.topMargin + 30,
                                   safe_arabic_text("التاريخ:          /  /     "))

    # Leave space beneath the header
    story.append(Spacer(1, 60))

    # Helper to build specific report tables
    def build_dogs_table(dogs):
        from reportlab.lib.units import mm
        header = [
            hdr('رقم'),
            hdr('اسم الكلب'),
            hdr('الكود'),
            hdr('السلالة'),
            hdr('الجنس'),
            hdr('الحالة'),
            hdr('الموقع'),
        ]
        
        rows = [header]
        for i, d in enumerate(dogs, start=1):
            gender_ar = 'ذكر' if d.gender.value == 'MALE' else 'أنثى'
            status_ar = {
                'ACTIVE': 'نشط', 'RETIRED': 'متقاعد', 
                'DECEASED': 'متوفى', 'TRAINING': 'تدريب'
            }.get(d.current_status.value, d.current_status.value)
            
            rows.append([
                para_ltr(i),
                para_ar(d.name or ''),
                para_ltr(d.code or ''),
                para_ar(d.breed or ''),
                para_ar(gender_ar),
                para_ar(status_ar),
                para_ar(d.location or ''),
            ])
        
        col_widths = [12*mm, 35*mm, 25*mm, 30*mm, 18*mm, 25*mm, 30*mm]
        table = Table(rows, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#603913')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#603913')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,1), (0,-1), 'LEFT'),  # رقم
            ('ALIGN', (2,1), (2,-1), 'LEFT'),  # الكود
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
        ]))
        return table
    
    def build_employees_table(employees):
        from reportlab.lib.units import mm
        header = [
            hdr('رقم'),
            hdr('الاسم'),
            hdr('الرقم الوظيفي'),
            hdr('الوظيفة'),
            hdr('تاريخ التعيين'),
            hdr('الحالة'),
            hdr('الهاتف'),
            hdr('البريد')
        ]
        
        rows = [header]
        role_map = {
            'مدرب': 'مدرب', 'سائس': 'معالج', 'طبيب': 'طبيب بيطري',
            'مسؤول مشروع': 'مدير', 'مربي': 'مربي'
        }
        
        for i, emp in enumerate(employees, start=1):
            status_ar = 'نشط' if emp.is_active else 'غير نشط'
            rows.append([
                para_ltr(i),
                para_ar(emp.name),
                para_ltr(emp.employee_id or ''),
                para_ar(role_map.get(emp.role.value, emp.role.value)),
                para_ltr(emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else ''),
                para_ar(status_ar),
                para_ltr(emp.phone or ''),
                para_ltr(emp.email or '')
            ])
        
        col_widths = [12*mm, 35*mm, 20*mm, 25*mm, 25*mm, 20*mm, 25*mm, 35*mm]
        table = Table(rows, repeatRows=1, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#854321')),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#854321')),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,1), (0,-1), 'LEFT'),  # رقم
            ('ALIGN', (2,1), (2,-1), 'LEFT'),  # الرقم الوظيفي
            ('ALIGN', (4,1), (4,-1), 'LEFT'),  # تاريخ التعيين
            ('ALIGN', (6,1), (6,-1), 'LEFT'),  # الهاتف
            ('ALIGN', (7,1), (7,-1), 'LEFT'),  # البريد
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
        ]))
        return table
    
    def build_table(data, header_color):
        """Build a generic table with Arabic/RTL support"""
        from reportlab.lib.units import mm
        
        # Convert first row (headers) to proper format
        if data and len(data) > 0:
            header_row = [hdr(cell) for cell in data[0]]
            data_rows = []
            
            # Convert data rows
            for row in data[1:]:
                converted_row = []
                for cell in row:
                    if isinstance(cell, str) and ARABIC_RE.search(cell):
                        converted_row.append(para_ar(cell))
                    else:
                        converted_row.append(para_ltr(str(cell)))
                data_rows.append(converted_row)
            
            # Combine header and data
            all_rows = [header_row] + data_rows
            
            # Calculate column widths (distribute evenly)
            num_cols = len(data[0]) if data else 1
            col_width = 190 / num_cols  # Total available width distributed
            col_widths = [col_width*mm] * num_cols
            
            table = Table(all_rows, repeatRows=1, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), header_color),
                ('GRID', (0,0), (-1,-1), 0.5, header_color),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
            ]))
            return table
        return None
    
    # Collect data based on report_type
    if report_type == 'dogs':
        story.append(Paragraph(shape_mixed("تقرير الكلاب"), title_style))
        # Fetch dogs accessible by the current user
        dogs = Dog.query.all() if user.role.value == 'GENERAL_ADMIN' \
            else Dog.query.filter_by(assigned_to_user_id=user.id).all()
        # Apply filters
        status_filter = filters.get('status')
        gender_filter = filters.get('gender')
        if status_filter:
            dogs = [d for d in dogs if d.current_status.value == status_filter]
        if gender_filter:
            dogs = [d for d in dogs if d.gender.value == gender_filter]
        story.append(build_dogs_table(dogs))
    elif report_type == 'employees':
        story.append(Paragraph(shape_mixed("تقرير الموظفين"), title_style))
        employees = Employee.query.all()
        role_filter = filters.get('role')
        status_filter = filters.get('status')
        if role_filter:
            employees = [e for e in employees if e.role.value == role_filter]
        if status_filter:
            is_active = (status_filter == 'ACTIVE')
            employees = [e for e in employees if e.is_active == is_active]
        story.append(build_employees_table(employees))
    
    elif report_type == 'training':
        story.append(Paragraph(shape_mixed("تقرير التدريب"), title_style))
        sessions = TrainingSession.query
        # Apply date range
        if start_date and end_date:
            sessions = sessions.filter(TrainingSession.session_date >= start_date,
                                       TrainingSession.session_date <= end_date)
        # Restrict to assigned dogs for project managers
        if user.role.value != 'GENERAL_ADMIN':
            assigned_ids = [d.id for d in Dog.query.filter_by(assigned_to_user_id=user.id).all()]
            sessions = sessions.filter(TrainingSession.dog_id.in_(assigned_ids))
        # Filter by category
        category_filter = filters.get('category')
        if category_filter:
            sessions = sessions.filter(TrainingSession.category == category_filter)
        sessions = sessions.all()
        story.append(build_training_table(sessions))
    elif report_type == 'veterinary':
        story.append(Paragraph(safe_arabic_text("تقرير الطبابة"), title_style))
        visits = VeterinaryVisit.query
        if start_date and end_date:
            visits = visits.filter(VeterinaryVisit.visit_date >= start_date,
                                   VeterinaryVisit.visit_date <= end_date)
        if user.role.value != 'GENERAL_ADMIN':
            assigned_ids = [d.id for d in Dog.query.filter_by(assigned_to_user_id=user.id).all()]
            visits = visits.filter(VeterinaryVisit.dog_id.in_(assigned_ids))
        visit_type_filter = filters.get('visit_type')
        if visit_type_filter:
            visits = visits.filter(VeterinaryVisit.visit_type == visit_type_filter)
        visits = visits.all()
        data = [['الكلب', 'الطبيب', 'نوع الزيارة', 'التاريخ', 'التشخيص', 'العلاج']]
        visit_type_map = {'ROUTINE': 'روتينية', 'EMERGENCY': 'طارئة', 'VACCINATION': 'تطعيم'}
        for v in visits:
            data.append([
                v.dog.name,
                v.vet.name,
                visit_type_map.get(v.visit_type.value, v.visit_type.value),
                v.visit_date.strftime('%Y-%m-%d'),
                v.diagnosis or '', v.treatment or ''
            ])
        story.append(build_table(data, colors.HexColor('#008080')))  # teal header
    elif report_type == 'production':
        story.append(Paragraph(safe_arabic_text("تقرير الإنتاج"), title_style))
        cycles = ProductionCycle.query
        if start_date and end_date:
            cycles = cycles.filter(ProductionCycle.mating_date >= start_date,
                                   ProductionCycle.mating_date <= end_date)
        cycle_type_filter = filters.get('cycle_type')
        if cycle_type_filter:
            cycles = cycles.filter(ProductionCycle.cycle_type == cycle_type_filter)
        cycles = cycles.all()
        data = [['الأم', 'الأب', 'نوع الدورة', 'تاريخ التزاوج',
                 'تاريخ الولادة المتوقع', 'تاريخ الولادة', 'النتيجة',
                 'عدد الجراء', 'الناجون']]
        cycle_map = {'NATURAL': 'طبيعي', 'ARTIFICIAL': 'صناعي'}
        result_map = {'SUCCESSFUL': 'ناجحة', 'FAILED': 'فاشلة', 'UNKNOWN': 'غير معروف'}
        for c in cycles:
            data.append([
                c.female.name if c.female else '',
                c.male.name if c.male else '',
                cycle_map.get(c.cycle_type.value, c.cycle_type.value),
                c.mating_date.strftime('%Y-%m-%d') if c.mating_date else '',
                c.expected_delivery_date.strftime('%Y-%m-%d') if c.expected_delivery_date else '',
                c.actual_delivery_date.strftime('%Y-%m-%d') if c.actual_delivery_date else '',
                result_map.get(c.result.value, '') if c.result else '',
                c.number_of_puppies or '',
                c.puppies_survived or ''
            ])
        story.append(build_table(data, colors.HexColor('#008000')))  # green header
    elif report_type == 'projects':
        story.append(Paragraph(safe_arabic_text("تقرير المشاريع"), title_style))
        projects = Project.query
        if start_date and end_date:
            projects = projects.filter(Project.start_date >= start_date,
                                       Project.start_date <= end_date)
        status_filter = filters.get('project_status')
        if status_filter:
            projects = projects.filter(Project.status == status_filter)
        projects = projects.all()
        data = [['اسم المشروع', 'الكود', 'الحالة', 'تاريخ البداية',
                 'تاريخ النهاية', 'المدير', 'الموقع']]
        status_map = {
            'ACTIVE': 'نشط', 'COMPLETED': 'منجز',
            'CANCELLED': 'ملغى', 'PLANNED': 'مخطط'
        }
        for p in projects:
            data.append([
                p.name, p.code or '',
                status_map.get(p.status.value, p.status.value),
                p.start_date.strftime('%Y-%m-%d') if p.start_date else '',
                p.end_date.strftime('%Y-%m-%d') if p.end_date else '',
                p.manager.full_name if p.manager else '',
                p.location or ''
            ])
        story.append(build_table(data, colors.HexColor('#7030A0')))  # purple header

    # Add generation timestamp and user
    story.append(Spacer(1, 20))
    story.append(Paragraph(safe_arabic_text(f"تم إنشاء التقرير في: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"), normal_style))
    story.append(Paragraph(safe_arabic_text(f"بواسطة: {user.full_name}"), normal_style))

    # Add notes and signature lines as seen in the DOCX templates
    story.append(Spacer(1, 40))
    story.append(Paragraph(safe_arabic_text("ملاحظات:"), header_style))
    story.append(Spacer(1, 60))
    story.append(Paragraph(safe_arabic_text("اسم المسؤول: ..............................     التوقيع: .............................."), normal_style))
    story.append(Paragraph(safe_arabic_text("اسم المدير: ..............................     التوقيع: .............................."), normal_style))
    
    # Build PDF with header
    doc.build(story, onFirstPage=build_header)

    # Log the report generation (with safe error handling)
    try:
        import uuid
        log_audit(user.id, 'EXPORT', 'Report', str(uuid.uuid4()), {
            'report_type': report_type,
            'filename': filename,
            'start_date': start_date.isoformat() if start_date else None,
            'end_date': end_date.isoformat() if end_date else None,
            'filters': filters
        })
    except Exception as e:
        # Don't let audit logging errors break report generation
        print(f"Warning: Could not log export audit trail: {e}")
    
    return filename

def generate_excel_report(report_type, start_date, end_date, user, filters=None):
    """
    Generate an Excel report for the specified type and filters
    :param report_type: one of 'dogs', 'employees', 'training', 'veterinary', 'breeding', 'projects'
    :param start_date: optional start date (datetime.date)
    :param end_date: optional end date (datetime.date)
    :param user: current_user for permission filtering
    :param filters: optional dict with keys appropriate to report_type
    :return: filename of generated Excel file
    """
    from k9.models.models import Dog, Employee, TrainingSession, VeterinaryVisit, ProductionCycle, Project
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from k9.utils.report_header import create_excel_header_data
    
    filters = filters or {}
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    
    # Get company header data
    header_data = create_excel_header_data()
    
    # Arabic font and styling
    arabic_font = Font(name='Calibri', size=11, bold=False)
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Generate filename
    filename = f"report_{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
    
    # Collect data based on report_type
    if report_type == 'dogs':
        ws.title = "تقرير الكلاب"
        
        # Headers
        headers = ['رقم', 'اسم الكلب', 'الكود', 'السلالة', 'الجنس', 'الحالة', 'الموقع', 'العمر']
        
        # Fetch dogs accessible by the current user
        dogs = Dog.query.all() if user.role.value == 'GENERAL_ADMIN' \
            else Dog.query.filter_by(assigned_to_user_id=user.id).all()
        
        # Apply filters
        status_filter = filters.get('status')
        gender_filter = filters.get('gender')
        if status_filter:
            dogs = [d for d in dogs if d.current_status.value == status_filter]
        if gender_filter:
            dogs = [d for d in dogs if d.gender.value == gender_filter]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Adjust data start row to account for header
        data_start_row = current_row + 1
        
        # Write data
        for row, dog in enumerate(dogs, data_start_row):
            gender_ar = 'ذكر' if dog.gender.value == 'MALE' else 'أنثى'
            status_ar = {
                'ACTIVE': 'نشط', 'RETIRED': 'متقاعد', 
                'DECEASED': 'متوفى', 'TRAINING': 'تدريب'
            }.get(dog.current_status.value, dog.current_status.value)
            
            age = (datetime.now().date() - dog.date_of_birth).days // 365 if dog.date_of_birth else ''
            
            data = [row-1, dog.name or '', dog.code or '', dog.breed or '', 
                   gender_ar, status_ar, dog.location or '', age]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = arabic_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
    
    elif report_type == 'employees':
        ws.title = "تقرير الموظفين"
        
        # Headers
        headers = ['رقم', 'الاسم', 'الرقم الوظيفي', 'الوظيفة', 'تاريخ التعيين', 'الحالة', 'الهاتف', 'البريد']
        
        employees = Employee.query.all()
        role_filter = filters.get('role')
        status_filter = filters.get('status')
        if role_filter:
            employees = [e for e in employees if e.role.value == role_filter]
        if status_filter:
            is_active = (status_filter == 'ACTIVE')
            employees = [e for e in employees if e.is_active == is_active]
        
        # Add company header (same logic as dogs section)
        current_row = 1
        try:
            if os.path.exists(header_data['company_logo_path']):
                img = XLImage(header_data['company_logo_path'])
                img.height = 60
                img.width = 120
                ws.add_image(img, f'D{current_row}')
        except:
            pass
        
        for i, info in enumerate(header_data['english_info']):
            ws.cell(row=current_row + i, column=1).value = info
            ws.cell(row=current_row + i, column=1).font = Font(name='Calibri', size=9)
        
        for i, info in enumerate(header_data['arabic_info']):
            ws.cell(row=current_row + i, column=8).value = info
            ws.cell(row=current_row + i, column=8).font = Font(name='Calibri', size=9)
            ws.cell(row=current_row + i, column=8).alignment = Alignment(horizontal='right')
        
        current_row += 6
        ws.cell(row=current_row, column=4).value = "تقرير الموظفين"
        ws.cell(row=current_row, column=4).font = Font(name='Calibri', size=14, bold=True)
        ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Write data
        role_map = {
            'TRAINER': 'مدرب', 'HANDLER': 'معالج', 'VET': 'طبيب بيطري',
            'PROJECT_MANAGER': 'مدير مشروع', 'BREEDER': 'مربي'
        }
        
        for row, emp in enumerate(employees, 2):
            status_ar = 'نشط' if emp.is_active else 'غير نشط'
            data = [row-1, emp.name, emp.employee_id or '', 
                   role_map.get(emp.role.value, emp.role.value),
                   emp.hire_date.strftime('%Y-%m-%d') if emp.hire_date else '',
                   status_ar, emp.phone or '', emp.email or '']
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = arabic_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
    
    elif report_type == 'training':
        ws.title = "تقرير التدريب"
        
        # Headers
        headers = ['رقم', 'اسم الكلب', 'المدرب', 'الفئة', 'الموضوع', 'التاريخ', 'المدة (دقيقة)', 'التقييم']
        
        sessions = TrainingSession.query
        # Apply date range
        if start_date and end_date:
            sessions = sessions.filter(TrainingSession.session_date >= start_date,
                                       TrainingSession.session_date <= end_date)
        # Restrict to assigned dogs for project managers
        if user.role.value != 'GENERAL_ADMIN':
            assigned_ids = [d.id for d in Dog.query.filter_by(assigned_to_user_id=user.id).all()]
            sessions = sessions.filter(TrainingSession.dog_id.in_(assigned_ids))
        # Filter by category
        category_filter = filters.get('category')
        if category_filter:
            sessions = sessions.filter(TrainingSession.category == category_filter)
        sessions = sessions.all()
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Write data
        category_map = {
            'OBEDIENCE': 'طاعة', 'DETECTION': 'كشف', 'AGILITY': 'رشاقة',
            'ATTACK': 'هجوم', 'FITNESS': 'لياقة'
        }
        
        for row, session in enumerate(sessions, 2):
            data = [row-1, 
                   session.dog.name if session.dog else '',
                   session.trainer.name if session.trainer else '',
                   category_map.get(session.category.value, session.category.value),
                   session.subject or '',
                   session.session_date.strftime('%Y-%m-%d') if session.session_date else '',
                   session.duration or '',
                   f"{session.success_rating}/10" if session.success_rating else '']
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.font = arabic_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save file
    wb.save(filepath)
    
    # Log the report generation (with safe error handling)
    try:
        import uuid
        log_audit(user.id, 'EXPORT', 'Report', str(uuid.uuid4()), {
            'report_type': report_type,
            'format': 'excel',
            'filename': filename,
            'start_date': start_date.isoformat() if start_date else None,
            'end_date': end_date.isoformat() if end_date else None,
            'filters': filters
        })
    except Exception as e:
        # Don't let audit logging errors break report generation
        print(f"Warning: Could not log export audit trail: {e}")
    
    return filename

def get_user_permissions(user):
    """Get user permissions based on role and SubPermission grants"""
    if user.role.value == 'GENERAL_ADMIN':
        return {
            'dogs': True,
            'employees': True,
            'training': True,
            'veterinary': True,
            'production': True,
            'projects': True,
            'attendance': True,
            'reports': True,
            'admin': True
        }
    else:  # PROJECT_MANAGER
        # Load permissions from SubPermission table for PROJECT_MANAGER
        from k9.models.models import SubPermission, Project
        permissions = {section: False for section in ['dogs', 'employees', 'training', 'veterinary', 'production', 'projects', 'attendance', 'reports', 'admin']}
        
        # Check if user has any granted permissions in SubPermission table
        user_permissions = SubPermission.query.filter_by(user_id=user.id, is_granted=True).all()
        
        # Map SubPermission sections to the main sections
        section_mapping = {
            'Dogs': 'dogs',
            'Employees': 'employees', 
            'Training': 'training',
            'Veterinary': 'veterinary',
            'Production': 'production',
            'Projects': 'projects',
            'Attendance': 'attendance',
            'Reports': 'reports',
            'Analytics': 'reports'
        }
        
        # If user has any granted permissions, enable those sections
        for perm in user_permissions:
            mapped_section = section_mapping.get(perm.section)
            if mapped_section and mapped_section in permissions:
                permissions[mapped_section] = True
                
        # Also check assigned projects - if user has projects assigned, they should see projects section
        user_projects = Project.query.filter_by(manager_id=user.id).all()
        if user_projects:
            permissions['projects'] = True
        
        return permissions

def get_project_manager_permissions(user, project_id):
    """Get granular permissions for a PROJECT_MANAGER user on a specific project"""
    from k9.models.models import ProjectManagerPermission, UserRole
    
    if user.role == UserRole.GENERAL_ADMIN:
        # GENERAL_ADMIN has all permissions
        return {
            'can_manage_assignments': True,
            'can_manage_shifts': True,
            'can_manage_attendance': True,
            'can_manage_training': True,
            'can_manage_incidents': True,
            'can_manage_performance': True,
            'can_view_veterinary': True,
            'can_view_production': True
        }
    
    if user.role == UserRole.PROJECT_MANAGER:
        # Check if permission record exists
        permission = ProjectManagerPermission.query.filter_by(
            user_id=user.id,
            project_id=project_id
        ).first()
        
        if permission:
            return {
                'can_manage_assignments': permission.can_manage_assignments,
                'can_manage_shifts': permission.can_manage_shifts,
                'can_manage_attendance': permission.can_manage_attendance,
                'can_manage_training': permission.can_manage_training,
                'can_manage_incidents': permission.can_manage_incidents,
                'can_manage_performance': permission.can_manage_performance,
                'can_view_veterinary': permission.can_view_veterinary,
                'can_view_production': permission.can_view_production
            }
        else:
            # Create default permissions for new PROJECT_MANAGER with all permissions enabled
            permission = ProjectManagerPermission()
            permission.user_id = user.id
            permission.project_id = project_id
            # Explicitly set all permissions to True (override any defaults)
            permission.can_manage_assignments = True
            permission.can_manage_shifts = True
            permission.can_manage_attendance = True
            permission.can_manage_training = True
            permission.can_manage_incidents = True
            permission.can_manage_performance = True
            permission.can_view_veterinary = True
            permission.can_view_production = True
            db.session.add(permission)
            db.session.commit()
            
            return {
                'can_manage_assignments': True,
                'can_manage_shifts': True,
                'can_manage_attendance': True,
                'can_manage_training': True,
                'can_manage_incidents': True,
                'can_manage_performance': True,
                'can_view_veterinary': True,
                'can_view_production': True
            }
    
    # Default: no permissions for other roles
    return {
        'can_manage_assignments': False,
        'can_manage_shifts': False,
        'can_manage_attendance': False,
        'can_manage_training': False,
        'can_manage_incidents': False,
        'can_manage_performance': False,
        'can_view_veterinary': False,
        'can_view_breeding': False
    }

def get_user_assigned_projects(user):
    """Get projects with data that PROJECT_MANAGER can access based on SubPermission"""
    from k9.models.models import Project, UserRole, SubPermission
    
    if user.role == UserRole.GENERAL_ADMIN:
        return Project.query.all()
    elif user.role == UserRole.PROJECT_MANAGER:
        # Get projects where this user is assigned as manager
        assigned_projects = Project.query.filter_by(manager_id=user.id).all()
        
        # Also get projects from SubPermission grants
        project_permissions = SubPermission.query.filter_by(
            user_id=user.id, 
            section="Projects",
            is_granted=True
        ).filter(SubPermission.project_id.isnot(None)).all()
        
        # Add projects from permission grants
        permission_project_ids = [p.project_id for p in project_permissions]
        permission_projects = Project.query.filter(Project.id.in_(permission_project_ids)).all() if permission_project_ids else []
        
        # Combine and deduplicate
        all_projects = list(set(assigned_projects + permission_projects))
        return all_projects
    
    return []

def get_user_accessible_dogs(user):
    """Get dogs that PROJECT_MANAGER can access based on SubPermission and project assignments"""
    from k9.models.models import Dog, UserRole, SubPermission, ProjectAssignment, Project, PermissionType
    
    if user.role == UserRole.GENERAL_ADMIN:
        return Dog.query.all()
    elif user.role == UserRole.PROJECT_MANAGER:
        # Check if user has VIEW permission for Dogs
        dogs_permission = SubPermission.query.filter_by(
            user_id=user.id,
            section="Dogs",
            subsection="عرض قائمة الكلاب",
            permission_type=PermissionType.VIEW,
            is_granted=True
        ).first()
        
        if not dogs_permission:
            return []
            
        # Optimized single query: Get all dogs assigned to user's projects
        user_project_ids = db.session.query(ProjectAssignment.project_id).filter_by(user_id=user.id).subquery()
        
        dogs = db.session.query(Dog).join(ProjectAssignment).filter(
            ProjectAssignment.project_id.in_(user_project_ids),
            ProjectAssignment.dog_id.isnot(None)
        ).distinct().all()
        
        return dogs
    
    return []

def get_user_accessible_employees(user):
    """Get employees that PROJECT_MANAGER can access based on SubPermission and project assignments"""
    from k9.models.models import Employee, UserRole, SubPermission, ProjectAssignment, Project, PermissionType
    
    if user.role == UserRole.GENERAL_ADMIN:
        return Employee.query.all()
    elif user.role == UserRole.PROJECT_MANAGER:
        # Check if user has VIEW permission for Employees
        employees_permission = SubPermission.query.filter_by(
            user_id=user.id,
            section="Employees",
            subsection="عرض قائمة الموظفين",
            permission_type=PermissionType.VIEW,
            is_granted=True
        ).first()
        
        if not employees_permission:
            return []
            
        # Optimized single query: Get all employees assigned to user's projects
        user_project_ids = db.session.query(ProjectAssignment.project_id).filter_by(user_id=user.id).subquery()
        
        employees = db.session.query(Employee).join(ProjectAssignment).filter(
            ProjectAssignment.project_id.in_(user_project_ids),
            ProjectAssignment.employee_id.isnot(None)
        ).distinct().all()
        
        return employees
    
    return []

def get_user_projects(user):
    """Get projects assigned to a PROJECT_MANAGER user"""
    from k9.models.models import Project, UserRole, Employee, EmployeeRole
    
    if user.role == UserRole.GENERAL_ADMIN:
        return Project.query.all()
    elif user.role == UserRole.PROJECT_MANAGER:
        # Get projects where this user is the manager through employee profile
        direct_projects = []
        
        # Also check through employee profile assignment
        employee = Employee.query.filter_by(user_account_id=user.id, role=EmployeeRole.PROJECT_MANAGER).first()
        if employee:
            employee_projects = employee.projects
            # Combine and deduplicate
            all_projects = list(set(direct_projects + employee_projects))
            return all_projects
        
        return direct_projects
    
    return []

def check_project_access(user, project_id):
    """Check if user has access to a specific project"""
    from k9.models.models import Project, UserRole
    
    if user.role == UserRole.GENERAL_ADMIN:
        return True
    elif user.role == UserRole.PROJECT_MANAGER:
        project = Project.query.get(project_id)
        # Check through employee profile
        employee = Employee.query.filter_by(user_account_id=user.id).first()
        return project and employee and project.project_manager_id == employee.id
    
    return False

def filter_data_by_project_access(user, query, model_class):
    """Filter query results based on user's project access"""
    from k9.models.models import UserRole
    
    if user.role == UserRole.GENERAL_ADMIN:
        return query
    elif user.role == UserRole.PROJECT_MANAGER:
        user_projects = get_user_projects(user)
        project_ids = [p.id for p in user_projects]
        
        # Apply filtering based on model type
        if hasattr(model_class, 'project_id'):
            return query.filter(model_class.project_id.in_(project_ids))
        elif hasattr(model_class, 'projects'):
            # For models with many-to-many relationship with projects
            return query.join(model_class.projects).filter(model_class.id.in_(project_ids))
    
    return query.filter(False)  # No access by default

def ensure_employee_user_linkage():
    """Ensure all PROJECT_MANAGER employees have linked user accounts"""
    from k9.models.models import Employee, User, EmployeeRole, UserRole
    from werkzeug.security import generate_password_hash
    import secrets
    
    # Find PROJECT_MANAGER employees without user accounts
    pm_employees = Employee.query.filter_by(role=EmployeeRole.PROJECT_MANAGER, user_account_id=None).all()
    
    created_users = []
    for employee in pm_employees:
        # Create user account for this employee
        username = f"pm_{employee.employee_id.lower()}"
        temp_password = secrets.token_urlsafe(12)
        
        new_user = User()
        new_user.username = username
        new_user.email = employee.email or f"{username}@k9ops.local"
        new_user.password_hash = generate_password_hash(temp_password)
        new_user.full_name = employee.name
        new_user.role = UserRole.PROJECT_MANAGER
        new_user.active = True
        
        db.session.add(new_user)
        db.session.flush()
        
        # Link employee to user
        employee.user_account_id = new_user.id
        
        created_users.append({
            'user': new_user,
            'employee': employee,
            'temp_password': temp_password
        })
    
    if created_users:
        db.session.commit()
    
    return created_users

def get_employee_profile_for_user(user):
    """Get the employee profile for a PROJECT_MANAGER user"""
    from k9.models.models import Employee, EmployeeRole
    
    return Employee.query.filter_by(user_account_id=user.id, role=EmployeeRole.PROJECT_MANAGER).first()

def get_user_active_projects(user):
    """Get active projects for a PROJECT_MANAGER user"""
    from k9.models.models import Project, UserRole, Employee, EmployeeRole, ProjectStatus
    
    if user.role == UserRole.GENERAL_ADMIN:
        return Project.query.filter_by(status=ProjectStatus.ACTIVE).all()
    elif user.role == UserRole.PROJECT_MANAGER:
        # Get active projects where this user is the manager through employee profile
        active_projects = []
        
        # Also check through employee profile assignment for active projects
        employee = Employee.query.filter_by(user_account_id=user.id, role=EmployeeRole.PROJECT_MANAGER).first()
        if employee:
            employee_active_projects = [p for p in employee.projects if p.status == ProjectStatus.ACTIVE]
            # Combine and deduplicate
            all_active_projects = list(set(active_projects + employee_active_projects))
            return all_active_projects
        
        return active_projects
    
    return []

def get_user_all_projects(user):
    """Get ALL projects (regardless of status) for a PROJECT_MANAGER user"""
    from k9.models.models import Project, UserRole, Employee, EmployeeRole
    
    if user.role == UserRole.GENERAL_ADMIN:
        return Project.query.all()
    elif user.role == UserRole.PROJECT_MANAGER:
        # Get all projects where this user is the manager through employee profile
        manager_projects = []
        
        # Also check through employee profile assignment
        employee = Employee.query.filter_by(user_account_id=user.id, role=EmployeeRole.PROJECT_MANAGER).first()
        if employee:
            employee_projects = employee.projects
            # Combine and deduplicate
            all_projects = list(set(manager_projects + employee_projects))
            return all_projects
        
        return manager_projects
    
    return []

def validate_project_manager_assignment(employee_id, project):
    """
    Validate that a project manager can be assigned to a project.
    Enforces the rule: One project manager can only have one active/planned project at a time.
    
    Args:
        employee_id: Employee ID to validate
        project: Project object (for updates - to exclude current project)
    
    Returns:
        tuple: (can_assign: bool, error_message: str)
    """
    from k9.models.models import Project, Employee, ProjectStatus
    
    # Check if this employee has any existing active/planned projects
    existing_projects_query = Project.query.filter(
        Project.project_manager_id == employee_id,
        Project.status.in_([ProjectStatus.PLANNED, ProjectStatus.ACTIVE])
    )
    
    # Exclude current project if we're updating
    if project:
        existing_projects_query = existing_projects_query.filter(Project.id != project.id)
    
    existing_project = existing_projects_query.first()
    
    if existing_project:
        employee = Employee.query.get(employee_id)
        employee_name = employee.name if employee else "غير معروف"
        return False, f'مسؤول المشروع {employee_name} لديه مشروع نشط بالفعل: {existing_project.name}. لا يمكن تعيين أكثر من مشروع واحد في نفس الوقت.'
    
    return True, None

def validate_project_status_change(project, new_status, current_user=None):
    """
    Validate if a project status change is allowed, especially for project manager constraints.
    
    Rules:
    - If changing to ACTIVE or PLANNED, ensure the project manager doesn't have other ongoing projects
    - If project manager is being changed while project is ACTIVE/PLANNED, validate the new manager
    """
    from k9.models.models import ProjectStatus, UserRole
    
    # If changing to ACTIVE or PLANNED status, validate project manager
    if new_status in [ProjectStatus.ACTIVE, ProjectStatus.PLANNED] and project.project_manager_id:
        # Get the project manager user
        from k9.models.models import Employee, EmployeeRole
        employee = Employee.query.get(project.project_manager_id)
        
        if employee and employee.role == EmployeeRole.PROJECT_MANAGER:
            # Temporarily set the new status for validation
            original_status = project.status
            project.status = new_status
            
            can_assign, error_msg = validate_project_manager_assignment(employee.id, project)
            
            # Restore original status
            project.status = original_status
            
            if not can_assign:
                return False, error_msg
    
    return True, None

def get_dog_active_project(dog_id, activity_date=None):
    """Get the active project for a dog at a specific date"""
    from k9.models.models import ProjectAssignment
    from datetime import datetime
    
    if activity_date is None:
        activity_date = datetime.utcnow()
    
    # Convert date to datetime if needed
    if hasattr(activity_date, 'date'):
        activity_datetime = activity_date
    else:
        activity_datetime = datetime.combine(activity_date, datetime.min.time())
    
    # Find active assignment for the dog at the given date
    active_assignment = ProjectAssignment.query.filter(
        ProjectAssignment.dog_id == str(dog_id),
        ProjectAssignment.assigned_from <= activity_datetime,
        db.or_(
            ProjectAssignment.assigned_to.is_(None),  # Still active
            ProjectAssignment.assigned_to >= activity_datetime  # Not yet ended
        )
    ).first()
    
    return active_assignment.project_id if active_assignment else None

def auto_link_dog_activity_to_project(dog_id, activity_date=None):
    """Automatically determine and return the project_id for a dog activity"""
    return get_dog_active_project(dog_id, activity_date)

def get_project_linked_activities(project_id, start_date=None, end_date=None):
    """Get all activities linked to a project within a date range"""
    from k9.models.models import TrainingSession, VeterinaryVisit, Incident, Suspicion, ProjectAssignment
    from datetime import datetime
    
    activities = {
        'training_sessions': [],
        'veterinary_visits': [],
        'incidents': [],
        'suspicions': []
    }
    
    # Get project assignments to filter activities by assignment period
    assignments = ProjectAssignment.query.filter_by(project_id=str(project_id)).all()
    dog_assignments = {a.dog_id: a for a in assignments if a.dog_id}
    
    # Training sessions
    training_query = TrainingSession.query.filter_by(project_id=str(project_id))
    if start_date:
        training_query = training_query.filter(TrainingSession.session_date >= start_date)
    if end_date:
        training_query = training_query.filter(TrainingSession.session_date <= end_date)
    
    for session in training_query.all():
        # Check if session is within dog's assignment period
        assignment = dog_assignments.get(session.dog_id)
        if assignment:
            session_date = session.session_date
            if assignment.assigned_from <= session_date:
                if not assignment.assigned_to or assignment.assigned_to >= session_date:
                    activities['training_sessions'].append(session)
    
    # Veterinary visits
    vet_query = VeterinaryVisit.query.filter_by(project_id=str(project_id))
    if start_date:
        vet_query = vet_query.filter(VeterinaryVisit.visit_date >= start_date)
    if end_date:
        vet_query = vet_query.filter(VeterinaryVisit.visit_date <= end_date)
    
    for visit in vet_query.all():
        # Check if visit is within dog's assignment period
        assignment = dog_assignments.get(visit.dog_id)
        if assignment:
            visit_date = visit.visit_date
            if assignment.assigned_from <= visit_date:
                if not assignment.assigned_to or assignment.assigned_to >= visit_date:
                    activities['veterinary_visits'].append(visit)
    
    # Incidents (already have project_id)
    incident_query = Incident.query.filter_by(project_id=str(project_id))
    if start_date:
        incident_query = incident_query.filter(Incident.incident_date >= start_date)
    if end_date:
        incident_query = incident_query.filter(Incident.incident_date <= end_date)
    activities['incidents'] = incident_query.all()
    
    # Suspicions (already have project_id)
    suspicion_query = Suspicion.query.filter_by(project_id=str(project_id))
    if start_date:
        suspicion_query = suspicion_query.filter(Suspicion.discovery_date >= start_date)
    if end_date:
        suspicion_query = suspicion_query.filter(Suspicion.discovery_date <= end_date)
    activities['suspicions'] = suspicion_query.all()
    
    return activities
